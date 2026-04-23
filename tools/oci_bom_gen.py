#!/usr/bin/env python3
"""
OCI Deal Accelerator — Bill of Materials (BOM) Generator

Generates a professional .xlsx BOM from a YAML spec, resolving SKUs
against the OCI SKU catalog. Only requested services appear in the output.

Usage:
    python oci_bom_gen.py --spec bom-spec.yaml --output customer-bom.xlsx
    python oci_bom_gen.py --spec bom-spec.yaml --output bom.xlsx --catalog kb/pricing/oci-sku-catalog.yaml
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ── Oracle Redwood palette ────────────────────────────────────────
class Colors:
    TEAL = "2C5967"
    TEAL_LIGHT = "E8F0F2"
    DARK_BG = "312D2A"
    COPPER = "AA643B"
    WHITE = "FFFFFF"
    GRAY_100 = "F5F4F2"
    GRAY_300 = "D4D1CC"
    GRAY_500 = "6F6964"
    GRAY_700 = "3A3631"
    BLACK = "000000"


# ── Column definitions ────────────────────────────────────────────
# Base columns (always present)
BASE_COLS = [
    ("Part (SKU)", 14),
    ("Product / Description", 55),
    ("Metric", 32),
    ("Unit Price (USD)", 16),
    ("Hours/Units", 12),
    ("Qty", 10),
    ("Months", 10),
    ("Discount", 10),
    ("Monthly USD\n(w/o discount)", 18),
    ("Monthly USD\n(w/ discount)", 18),
]

# Conversion columns (added when conversion is enabled)
# Placeholders — header text is set dynamically with target currency
CONV_COLS = [
    ("Monthly {cur}\n(w/o tax)", 18),
    ("Monthly {cur}\n(w/ tax)", 18),
]

# Always-last column
TAIL_COLS = [
    ("Cost %", 10),
]


def pick(mapping: dict, *keys, default=""):
    """Return the first non-empty value for any of the provided keys."""
    if not isinstance(mapping, dict):
        return default
    for key in keys:
        if key in mapping:
            value = mapping.get(key)
            if value is not None:
                return value
    return default


class OCIBomGenerator:
    """Generate OCI Bill of Materials (.xlsx) with Oracle Redwood styling."""

    def __init__(
        self,
        customer: str = "",
        project: str = "",
        prepared_by: str = "",
        bom_date: str = "",
        reference_label: str = "",
        currency: str = "USD",
        realm: str = "",
        service_type: str = "",
        conversion: Optional[dict] = None,
        notes: Optional[list] = None,
    ):
        self.customer = customer
        self.project = project
        self.prepared_by = prepared_by
        self.bom_date = bom_date or date.today().isoformat()
        self.reference_label = reference_label
        self.currency = currency
        self.realm = realm
        self.service_type = service_type
        self.conversion = conversion or {"enabled": False}
        self.notes = notes or []

        self.catalog: dict = {}          # sku_code -> {product, metric, ...}
        self.catalog_meta: dict = {}     # top-level catalog metadata
        self.category_order: list = []   # ordered category keys
        self.category_names: dict = {}   # cat_key -> display_name
        self.line_items: list = []       # resolved line items
        self.wb = Workbook()

    # ── Catalog loading ───────────────────────────────────────────

    def load_catalog(self, catalog_path: Optional[str] = None):
        """Load SKU catalog from YAML. Auto-discovers path if not given."""
        if not catalog_path:
            # Auto-discover relative to this script
            tools_dir = Path(__file__).resolve().parent
            catalog_path = str(tools_dir.parent / "kb" / "pricing" / "oci-sku-catalog.yaml")

        with open(catalog_path, "r") as f:
            raw = yaml.safe_load(f)

        self.catalog_meta = {
            k: v for k, v in raw.items() if k != "categories"
        }

        categories = raw.get("categories", {})
        for cat_key, cat_data in categories.items():
            self.category_order.append(cat_key)
            self.category_names[cat_key] = cat_data.get("display_name", cat_key)
            for sku_entry in cat_data.get("skus", []):
                sku_code = str(sku_entry["sku"])
                self.catalog[sku_code] = {
                    **sku_entry,
                    "sku": sku_code,
                    "category": cat_key,
                }

    # ── Line items ────────────────────────────────────────────────

    def add_line_item(
        self,
        sku: str,
        qty: float = 0,
        hours_units: Optional[float] = None,
        months: int = 12,
        discount: float = 0.0,
        custom_label: str = "",
        custom_note: str = "",
    ):
        """Add a line item, resolving SKU details from catalog."""
        sku = str(sku)
        cat_entry = self.catalog.get(sku, {})

        note = custom_note
        if not cat_entry:
            warn = f"[WARN] SKU {sku!r} not in catalog — included as line item with $0 price and warning note"
            print(warn, file=sys.stderr)
            flag = "⚠ SKU not in catalog — estimated $0, must be confirmed before quoting"
            note = f"{note}. {flag}" if note else flag
        elif cat_entry.get("estimate"):
            flag = "⚠ Estimate only — billed via underlying compute/storage; confirm actual shape + hours"
            note = f"{note}. {flag}" if note else flag

        item = {
            "sku": sku,
            "product": cat_entry.get("product", f"Unknown SKU ({sku})"),
            "metric": cat_entry.get("metric", ""),
            "list_price_usd": cat_entry.get("list_price_usd", 0),
            "hours_units": hours_units if hours_units is not None else cat_entry.get("default_hours_units", 730),
            "qty": qty,
            "months": months,
            "discount": discount,
            "category": cat_entry.get("category", "other"),
            "discountable": cat_entry.get("discountable", True),
            "billing_type": cat_entry.get("billing_type", "hourly"),
            "custom_label": custom_label,
            "custom_note": note,
        }
        self.line_items.append(item)

    # ── Excel generation ──────────────────────────────────────────

    def _col_defs(self) -> List[Tuple[str, int]]:
        """Return full column definitions based on conversion setting."""
        cols = list(BASE_COLS)
        if self.conversion.get("enabled"):
            cur = self.conversion.get("target_currency", "BRL")
            cols += [(h.format(cur=cur), w) for h, w in CONV_COLS]
        cols += list(TAIL_COLS)
        return cols

    def _col_count(self) -> int:
        return len(self._col_defs())

    @staticmethod
    def _monthly_values(item: dict) -> tuple[float, float]:
        """Return monthly values with and without discount for one BOM line."""
        price = float(item.get("list_price_usd", 0) or 0)
        hours_units = float(item.get("hours_units", 0) or 0)
        qty = float(item.get("qty", 0) or 0)
        discount = float(item.get("discount", 0) or 0)
        monthly_wo = price * hours_units * qty
        monthly_w = monthly_wo * (1 - discount)
        return monthly_wo, monthly_w

    def _build_sheet(self, ws):
        """Build the complete BOM sheet."""
        ws.title = "BOM"
        col_defs = self._col_defs()
        num_cols = len(col_defs)
        conv = self.conversion.get("enabled", False)

        # ── Styles ────────────────────────────────────────────────
        thin_border = Border(
            left=Side(style="thin", color=Colors.GRAY_300),
            right=Side(style="thin", color=Colors.GRAY_300),
            top=Side(style="thin", color=Colors.GRAY_300),
            bottom=Side(style="thin", color=Colors.GRAY_300),
        )
        header_font = Font(name="Segoe UI", size=11, bold=True, color=Colors.WHITE)
        header_fill = PatternFill(start_color=Colors.TEAL, end_color=Colors.TEAL, fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cat_font = Font(name="Segoe UI", size=10, bold=True, color=Colors.TEAL)
        cat_fill = PatternFill(start_color=Colors.TEAL_LIGHT, end_color=Colors.TEAL_LIGHT, fill_type="solid")

        data_font = Font(name="Segoe UI", size=10, color=Colors.GRAY_700)
        alt_fill = PatternFill(start_color=Colors.GRAY_100, end_color=Colors.GRAY_100, fill_type="solid")

        total_font = Font(name="Segoe UI", size=11, bold=True, color=Colors.DARK_BG)
        total_fill = PatternFill(start_color=Colors.GRAY_300, end_color=Colors.GRAY_300, fill_type="solid")

        subtotal_font = Font(name="Segoe UI", size=10, bold=True, color=Colors.GRAY_700)

        # ── Header area ───────────────────────────────────────────
        row = 1
        title_font = Font(name="Segoe UI", size=16, bold=True, color=Colors.TEAL)
        ws.cell(row=row, column=1, value="Oracle Investment Proposal").font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        row += 1

        subtitle_font = Font(name="Segoe UI", size=10, color=Colors.GRAY_500)
        ws.cell(row=row, column=1, value=f"as of {self.bom_date}").font = subtitle_font
        row += 1

        meta_font = Font(name="Segoe UI", size=10, color=Colors.GRAY_700)
        meta_items = [
            ("Customer", self.customer),
            ("Project", self.project),
            ("Prepared by", self.prepared_by),
            ("Reference", self.reference_label),
            ("Currency", self.currency),
            ("Realm", self.realm),
            ("Service Type", self.service_type),
        ]
        for label, value in meta_items:
            if value:
                ws.cell(row=row, column=1, value=f"{label}:").font = Font(
                    name="Segoe UI", size=10, bold=True, color=Colors.GRAY_700
                )
                ws.cell(row=row, column=2, value=value).font = meta_font
                row += 1

        if conv:
            xr = self.conversion.get("exchange_rate", 1)
            tc = self.conversion.get("target_currency", "BRL")
            tr = self.conversion.get("tax_rate", 0)
            ws.cell(row=row, column=1, value="Exchange Rate:").font = Font(
                name="Segoe UI", size=10, bold=True, color=Colors.GRAY_700
            )
            ws.cell(row=row, column=2, value=f"1 USD = {xr} {tc}").font = meta_font
            row += 1
            if tr > 0:
                ws.cell(row=row, column=1, value="Tax Rate:").font = Font(
                    name="Segoe UI", size=10, bold=True, color=Colors.GRAY_700
                )
                ws.cell(row=row, column=2, value=f"{tr*100:.2f}%").font = meta_font
                row += 1

        row += 1  # blank row before table

        # ── Column headers ────────────────────────────────────────
        header_row = row
        for ci, (col_name, col_width) in enumerate(col_defs, 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = col_width
        row += 1

        # ── Group items by category ───────────────────────────────
        items_by_cat: dict[str, list] = {}
        for item in self.line_items:
            cat = item["category"]
            items_by_cat.setdefault(cat, []).append(item)

        # Column index references (1-based)
        COL_SKU = 1
        COL_PRODUCT = 2
        COL_METRIC = 3
        COL_PRICE = 4
        COL_HOURS = 5
        COL_QTY = 6
        COL_MONTHS = 7
        COL_DISC = 8
        COL_M_WO = 9    # monthly without discount
        COL_M_W = 10     # monthly with discount
        if conv:
            COL_CONV_WO = 11  # monthly converted without tax
            COL_CONV_W = 12   # monthly converted with tax
            COL_PCT = 13      # cost proportion
        else:
            COL_PCT = 11

        data_start_row = row
        cat_subtotal_rows = []

        # Ensure categories present in items but missing from catalog order
        # (e.g., "other" from unknown SKUs) still render at the end.
        render_order = list(self.category_order)
        for cat_key in items_by_cat:
            if cat_key not in render_order:
                render_order.append(cat_key)

        grand_total_wo = 0.0
        grand_total_w = 0.0
        grand_total_conv_wo = 0.0
        grand_total_conv_w = 0.0

        for cat_key in render_order:
            if cat_key not in items_by_cat:
                continue

            cat_items = items_by_cat[cat_key]
            display_name = self.category_names.get(
                cat_key,
                "Uncategorized — confirm SKUs" if cat_key == "other" else cat_key,
            )

            # Category header row
            ws.cell(row=row, column=COL_PRODUCT, value=display_name).font = cat_font
            for ci in range(1, num_cols + 1):
                ws.cell(row=row, column=ci).fill = cat_fill
                ws.cell(row=row, column=ci).border = thin_border
            cat_header_row = row
            row += 1

            cat_first_data = row
            cat_total_wo = 0.0
            cat_total_w = 0.0
            cat_total_conv_wo = 0.0
            cat_total_conv_w = 0.0

            for idx, item in enumerate(cat_items):
                # Static data
                ws.cell(row=row, column=COL_SKU, value=item["sku"])
                ws.cell(row=row, column=COL_PRODUCT, value=item["product"])
                ws.cell(row=row, column=COL_METRIC, value=item["metric"])
                ws.cell(row=row, column=COL_PRICE, value=item["list_price_usd"])
                ws.cell(row=row, column=COL_HOURS, value=item["hours_units"])
                ws.cell(row=row, column=COL_QTY, value=item["qty"])
                ws.cell(row=row, column=COL_MONTHS, value=item["months"])
                ws.cell(row=row, column=COL_DISC, value=item["discount"])

                monthly_wo, monthly_w = self._monthly_values(item)
                ws.cell(row=row, column=COL_M_WO, value=monthly_wo)
                ws.cell(row=row, column=COL_M_W, value=monthly_w)
                cat_total_wo += monthly_wo
                cat_total_w += monthly_w

                # Conversion formulas
                if conv:
                    xr = self.conversion.get("exchange_rate", 1)
                    tr = self.conversion.get("tax_rate", 0)
                    conv_wo = monthly_w * xr
                    conv_w = conv_wo * (1 + tr)
                    ws.cell(row=row, column=COL_CONV_WO, value=conv_wo)
                    ws.cell(row=row, column=COL_CONV_W, value=conv_w)
                    cat_total_conv_wo += conv_wo
                    cat_total_conv_w += conv_w

                # Styling per row
                is_alt = idx % 2 == 1
                for ci in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.font = data_font
                    cell.border = thin_border
                    if is_alt:
                        cell.fill = alt_fill
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=(ci == COL_PRODUCT),
                    )

                # Number formats
                ws.cell(row=row, column=COL_PRICE).number_format = '$#,##0.0000'
                ws.cell(row=row, column=COL_HOURS).number_format = '#,##0'
                ws.cell(row=row, column=COL_QTY).number_format = '#,##0.##'
                ws.cell(row=row, column=COL_MONTHS).number_format = '#,##0'
                ws.cell(row=row, column=COL_DISC).number_format = '0%'
                ws.cell(row=row, column=COL_M_WO).number_format = '$#,##0.00'
                ws.cell(row=row, column=COL_M_W).number_format = '$#,##0.00'
                if conv:
                    ws.cell(row=row, column=COL_CONV_WO).number_format = '#,##0.00'
                    ws.cell(row=row, column=COL_CONV_W).number_format = '#,##0.00'

                row += 1

            # Category subtotal row
            cat_last_data = row - 1
            ws.cell(row=row, column=COL_PRODUCT, value=f"Subtotal — {display_name}")
            ws.cell(row=row, column=COL_M_WO, value=cat_total_wo)
            ws.cell(row=row, column=COL_M_W, value=cat_total_w)
            ws.cell(row=row, column=COL_M_WO).number_format = '$#,##0.00'
            ws.cell(row=row, column=COL_M_W).number_format = '$#,##0.00'
            grand_total_wo += cat_total_wo
            grand_total_w += cat_total_w

            if conv:
                ws.cell(row=row, column=COL_CONV_WO, value=cat_total_conv_wo)
                ws.cell(row=row, column=COL_CONV_W, value=cat_total_conv_w)
                ws.cell(row=row, column=COL_CONV_WO).number_format = '#,##0.00'
                ws.cell(row=row, column=COL_CONV_W).number_format = '#,##0.00'
                grand_total_conv_wo += cat_total_conv_wo
                grand_total_conv_w += cat_total_conv_w

            for ci in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=ci)
                cell.font = subtotal_font
                cell.border = thin_border
                cell.fill = PatternFill(start_color="EDE9E5", end_color="EDE9E5", fill_type="solid")

            cat_subtotal_rows.append(row)

            row += 1  # blank row between categories

        # ── Grand total ───────────────────────────────────────────
        total_row = row
        ws.cell(row=row, column=COL_PRODUCT, value="TOTAL").font = total_font

        ws.cell(row=row, column=COL_M_WO, value=grand_total_wo)
        ws.cell(row=row, column=COL_M_W, value=grand_total_w)
        ws.cell(row=row, column=COL_M_WO).number_format = '$#,##0.00'
        ws.cell(row=row, column=COL_M_W).number_format = '$#,##0.00'

        if conv:
            ws.cell(row=row, column=COL_CONV_WO, value=grand_total_conv_wo)
            ws.cell(row=row, column=COL_CONV_W, value=grand_total_conv_w)
            ws.cell(row=row, column=COL_CONV_WO).number_format = '#,##0.00'
            ws.cell(row=row, column=COL_CONV_W).number_format = '#,##0.00'

        for ci in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=ci)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
        row += 1

        # ── ARR (Annual Run Rate) ────────────────────────────────
        arr_row = row
        ws.cell(row=row, column=COL_PRODUCT, value="Annual Run Rate (ARR)").font = Font(
            name="Segoe UI", size=11, bold=True, color=Colors.COPPER
        )
        ws.cell(row=row, column=COL_M_W, value=grand_total_w * 12)
        ws.cell(row=row, column=COL_M_W).number_format = '$#,##0.00'
        ws.cell(row=row, column=COL_M_W).font = Font(
            name="Segoe UI", size=11, bold=True, color=Colors.COPPER
        )
        if conv:
            ws.cell(row=row, column=COL_CONV_W, value=grand_total_conv_w * 12)
            ws.cell(row=row, column=COL_CONV_W).number_format = '#,##0.00'
            ws.cell(row=row, column=COL_CONV_W).font = Font(
                name="Segoe UI", size=11, bold=True, color=Colors.COPPER
            )
        for ci in range(1, num_cols + 1):
            ws.cell(row=row, column=ci).border = thin_border
        row += 2

        # ── Cost proportion formulas (back-fill) ─────────────────
        # Cost % = line monthly w/ discount / total monthly w/ discount
        for item_row in range(data_start_row, total_row):
            cell_val = ws.cell(row=item_row, column=COL_M_W).value
            if cell_val and str(cell_val).startswith("=") and "SUM" not in str(cell_val):
                mw_l = get_column_letter(COL_M_W)
                ws.cell(
                    row=item_row, column=COL_PCT,
                    value=f"=IF({mw_l}{total_row}=0,0,{mw_l}{item_row}/{mw_l}{total_row})",
                )
                ws.cell(row=item_row, column=COL_PCT).number_format = '0.0%'
                ws.cell(row=item_row, column=COL_PCT).font = data_font
                ws.cell(row=item_row, column=COL_PCT).border = thin_border

        # ── Notes ─────────────────────────────────────────────────
        if self.notes:
            ws.cell(row=row, column=1, value="Notes:").font = Font(
                name="Segoe UI", size=10, bold=True, color=Colors.GRAY_500
            )
            row += 1
            for note in self.notes:
                ws.cell(row=row, column=1, value=f"  - {note}").font = Font(
                    name="Segoe UI", size=9, color=Colors.GRAY_500
                )
                row += 1
            row += 1

        # ── Disclaimer ────────────────────────────────────────────
        disclaimer = self.catalog_meta.get("disclaimer", "")
        if disclaimer:
            ws.cell(row=row, column=1, value="Disclaimer:").font = Font(
                name="Segoe UI", size=8, bold=True, color=Colors.GRAY_500
            )
            row += 1
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=num_cols,
            )
            cell = ws.cell(row=row, column=1, value=disclaimer.strip())
            cell.font = Font(name="Segoe UI", size=8, color=Colors.GRAY_500)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 60

        # ── Print setup ───────────────────────────────────────────
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"{header_row}:{header_row}"

    # ── AppCA Export ──────────────────────────────────────────────

    def _build_appca_sheets(self):
        """Build the two sheets needed for AppCA import: 'Export to AppCA' and 'BOM.C1'."""

        # ── Styles ────────────────────────────────────────────────
        thin_border = Border(
            left=Side(style="thin", color=Colors.GRAY_300),
            right=Side(style="thin", color=Colors.GRAY_300),
            top=Side(style="thin", color=Colors.GRAY_300),
            bottom=Side(style="thin", color=Colors.GRAY_300),
        )
        header_font = Font(name="Segoe UI", size=11, bold=True, color=Colors.WHITE)
        header_fill = PatternFill(start_color=Colors.TEAL, end_color=Colors.TEAL, fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="Segoe UI", size=10, color=Colors.GRAY_700)
        cat_font = Font(name="Segoe UI", size=10, bold=True, color=Colors.TEAL)
        cat_fill = PatternFill(start_color=Colors.TEAL_LIGHT, end_color=Colors.TEAL_LIGHT, fill_type="solid")

        # ── Sheet: BOM.C1 ─────────────────────────────────────────
        ws_bom = self.wb.create_sheet("BOM.C1")

        # Row 1: global parameters
        ws_bom.cell(row=1, column=6, value="Hours: ").font = Font(name="Segoe UI", size=10, bold=True)
        ws_bom.cell(row=1, column=7, value=730)
        ws_bom.cell(row=1, column=8, value="Months: ").font = Font(name="Segoe UI", size=10, bold=True)
        months_val = self.line_items[0]["months"] if self.line_items else 12
        ws_bom.cell(row=1, column=9, value=months_val)

        # Row 4: title
        ws_bom.cell(row=4, column=4, value="Oracle Cloud Infrastructure - Bill of Material (BOM)").font = Font(
            name="Segoe UI", size=12, bold=True, color=Colors.TEAL
        )

        # Row 5: column headers (matching original BOM.C1 layout: C=SKU, D=Product, etc.)
        bom_headers = [
            (3, "SKU", 14), (4, "Product", 50), (5, "Metric", 30),
            (6, "Price List US$", 14), (7, "Hours/Units", 12),
            (8, "Qty", 10), (9, "Months", 10), (10, "Discount", 10),
            (11, "US$\n(w/o discount)", 16), (12, "US$\n(w/ discount)", 16),
        ]
        for col, name, width in bom_headers:
            cell = ws_bom.cell(row=5, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws_bom.column_dimensions[get_column_letter(col)].width = width

        # Row 6+: data rows grouped by category
        row = 6
        items_by_cat = {}
        for item in self.line_items:
            cat = item["category"]
            items_by_cat.setdefault(cat, []).append(item)

        render_order = list(self.category_order)
        for cat_key in items_by_cat:
            if cat_key not in render_order:
                render_order.append(cat_key)

        for cat_key in render_order:
            if cat_key not in items_by_cat:
                continue

            display_name = self.category_names.get(
                cat_key,
                "Uncategorized — confirm SKUs" if cat_key == "other" else cat_key,
            )

            # Category header
            ws_bom.cell(row=row, column=4, value=display_name).font = cat_font
            for ci in range(3, 13):
                ws_bom.cell(row=row, column=ci).fill = cat_fill
                ws_bom.cell(row=row, column=ci).border = thin_border
            row += 1

            for item in items_by_cat[cat_key]:
                ws_bom.cell(row=row, column=3, value=item["sku"])
                ws_bom.cell(row=row, column=4, value=item["product"])
                ws_bom.cell(row=row, column=5, value=item["metric"])
                ws_bom.cell(row=row, column=6, value=item["list_price_usd"])
                ws_bom.cell(row=row, column=7, value=item["hours_units"])
                ws_bom.cell(row=row, column=8, value=item["qty"])
                ws_bom.cell(row=row, column=9, value=item["months"])
                ws_bom.cell(row=row, column=10, value=item["discount"])

                monthly_wo, monthly_w = self._monthly_values(item)
                ws_bom.cell(row=row, column=11, value=monthly_wo)
                ws_bom.cell(row=row, column=12, value=monthly_w)

                # Formats
                ws_bom.cell(row=row, column=6).number_format = '#,##0.0000'
                ws_bom.cell(row=row, column=10).number_format = '0.0%'
                ws_bom.cell(row=row, column=11).number_format = '$#,##0.00'
                ws_bom.cell(row=row, column=12).number_format = '$#,##0.00'

                for ci in range(3, 13):
                    cell = ws_bom.cell(row=row, column=ci)
                    cell.font = data_font
                    cell.border = thin_border
                row += 1

        # ── Sheet: Export to AppCA ─────────────────────────────────
        ws_appca = self.wb.create_sheet("Export to AppCA", 0)  # first sheet

        # Title
        ws_appca.cell(row=1, column=1, value="TEMPLATE BOM - APPCA").font = Font(
            name="Segoe UI", size=14, bold=True, color=Colors.TEAL
        )
        ws_appca.merge_cells("A1:G1")

        # Headers — AppCA expects: SKU, QTY, HOURS, MONTHS, DISCOUNT, BURSTABLE, STATUS
        appca_headers = ["SKU", "QTY", "HOURS", "MONTHS", "DISCOUNT", "BURSTABLE", "STATUS"]
        appca_widths = [14, 10, 12, 10, 12, 10, 12]
        for ci, (name, width) in enumerate(zip(appca_headers, appca_widths), 1):
            cell = ws_appca.cell(row=2, column=ci, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws_appca.column_dimensions[get_column_letter(ci)].width = width

        # Data rows: only items with qty > 0
        appca_row = 3
        for item in self.line_items:
            if item["qty"] > 0:
                ws_appca.cell(row=appca_row, column=1, value=item["sku"])
                ws_appca.cell(row=appca_row, column=2, value=item["qty"])
                ws_appca.cell(row=appca_row, column=3, value=item["hours_units"])
                ws_appca.cell(row=appca_row, column=4, value=item["months"])
                ws_appca.cell(row=appca_row, column=5, value=item["discount"])
                ws_appca.cell(row=appca_row, column=6, value="No")    # BURSTABLE: No unless specified
                ws_appca.cell(row=appca_row, column=7, value="Yes")   # STATUS: required by AppCA

                ws_appca.cell(row=appca_row, column=5).number_format = '0.0%'

                for ci in range(1, 8):
                    cell = ws_appca.cell(row=appca_row, column=ci)
                    cell.font = data_font
                    cell.border = thin_border
                appca_row += 1

    # ── Public API ────────────────────────────────────────────────

    def save(self, filepath: str, appca: bool = False):
        """Build and save the workbook to an .xlsx file."""
        if appca:
            # Remove default empty sheet, build AppCA sheets
            default_sheet = self.wb.active
            self._build_appca_sheets()
            self.wb.remove(default_sheet)
        else:
            ws = self.wb.active
            self._build_sheet(ws)
        self.wb.save(filepath)
        print("  BOM saved: {}".format(filepath))

    @classmethod
    def from_spec(cls, spec: dict, catalog_path: Optional[str] = None) -> "OCIBomGenerator":
        """Build BOM from a YAML specification dict."""
        bom = spec.get("bom", spec)
        conversion = bom.get("conversion")
        if isinstance(conversion, dict):
            conversion = dict(conversion)
            if "target_currency" not in conversion and "currency" in conversion:
                conversion["target_currency"] = conversion.get("currency")

        gen = cls(
            customer=pick(bom, "customer_name", "customer"),
            project=pick(bom, "project_name", "project"),
            prepared_by=pick(bom, "prepared_by", "author"),
            bom_date=bom.get("date", ""),
            reference_label=pick(bom, "reference_label", "reference"),
            currency=bom.get("currency", "USD"),
            realm=bom.get("realm", ""),
            service_type=bom.get("service_type", ""),
            conversion=conversion,
            notes=bom.get("notes"),
        )

        gen.load_catalog(catalog_path)

        # Global discount from metadata — used as fallback for line items without explicit discount
        bom_meta = bom.get("metadata", {})
        global_discount = bom_meta.get("discount_pct", bom.get("discount_pct", 0.0))

        for item in bom.get("line_items", []):
            sku = pick(item, "sku", "part_number")
            if sku in ("", None):
                raise ValueError("BOM line item requires 'sku' or 'part_number'")
            gen.add_line_item(
                sku=str(sku),
                qty=pick(item, "qty", "quantity", default=0),
                hours_units=pick(item, "hours_units", "units", "hours", default=None),
                months=item.get("months", 12),
                discount=pick(item, "discount", "discount_pct", default=global_discount),
                custom_label=pick(item, "custom_label", "label"),
                custom_note=pick(item, "custom_note", "note", "notes"),
            )

        return gen


def main():
    parser = argparse.ArgumentParser(
        description="OCI Deal Accelerator — BOM Generator (.xlsx)"
    )
    parser.add_argument(
        "--spec", required=True,
        help="Path to BOM spec YAML file",
    )
    parser.add_argument(
        "--output", default="bom.xlsx",
        help="Output .xlsx file path (default: bom.xlsx)",
    )
    parser.add_argument(
        "--catalog", default=None,
        help="Override path to SKU catalog YAML (default: auto-discover)",
    )
    parser.add_argument(
        "--appca", action="store_true",
        help="Generate AppCA import format (2 sheets: 'Export to AppCA' + 'BOM.C1')",
    )
    args = parser.parse_args()

    with open(args.spec, "r") as f:
        spec = yaml.safe_load(f)

    gen = OCIBomGenerator.from_spec(spec, catalog_path=args.catalog)
    gen.save(args.output, appca=args.appca)


if __name__ == "__main__":
    main()
